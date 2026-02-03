"""
Excel Importer
Import configuration data from Excel files (.xlsx, .xls)
"""
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional

from .base_importer import BaseImporter, ImportResult


class ExcelImporter(BaseImporter):
    """Importer for Excel files (requires openpyxl)"""

    def __init__(self):
        self._openpyxl_available = None

    def _check_openpyxl(self) -> bool:
        """Check if openpyxl is available"""
        if self._openpyxl_available is None:
            try:
                import openpyxl
                self._openpyxl_available = True
            except ImportError:
                self._openpyxl_available = False
        return self._openpyxl_available

    @property
    def supported_extensions(self) -> List[str]:
        return ['.xlsx', '.xlsm']

    @property
    def format_name(self) -> str:
        return "Excel Workbook (.xlsx)"

    def validate_file(self, path: Path) -> Tuple[bool, Optional[str]]:
        """Validate Excel file"""
        if not self._check_openpyxl():
            return False, "Excel support requires openpyxl library. Install with: pip install openpyxl"

        if not path.exists():
            return False, f"File not found: {path}"

        if path.suffix.lower() not in self.supported_extensions:
            return False, f"Unsupported file type: {path.suffix}. Expected: {self.supported_extensions}"

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            # Check if there's at least a header row
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()
            if not first_row or all(v is None for v in first_row):
                return False, "Worksheet is empty or has no header row"
            return True, None
        except Exception as e:
            return False, f"Cannot read Excel file: {e}"

    def get_columns(self, path: Path) -> List[str]:
        """Get column names from first row"""
        if not self._check_openpyxl():
            return []

        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active

        columns = []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        for cell in first_row:
            if cell is not None:
                columns.append(str(cell).strip())
            else:
                columns.append(f"Column_{len(columns)+1}")

        wb.close()
        return columns

    def preview(self, path: Path, limit: int = 10) -> List[Dict[str, Any]]:
        """Preview first N records from Excel"""
        if not self._check_openpyxl():
            return []

        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active

        records = []
        headers = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                # Header row
                headers = [str(cell).strip() if cell else f"Column_{i+1}"
                          for i, cell in enumerate(row)]
            else:
                if row_idx > limit:
                    break
                # Data row
                record = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = str(cell).strip() if cell is not None else ''
                records.append(record)

        wb.close()
        return records

    def import_data(self, path: Path, column_mapping: Dict[str, str]) -> ImportResult:
        """Import data from Excel with column mapping"""
        result = ImportResult(success=False)

        if not self._check_openpyxl():
            result.errors.append("openpyxl library not available")
            return result

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active

            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell).strip() if cell else f"Column_{i+1}"
                              for i, cell in enumerate(row)]
                    continue

                try:
                    row_data = {headers[i]: str(cell).strip() if cell is not None else ''
                               for i, cell in enumerate(row) if i < len(headers)}

                    # Apply column mapping
                    mapped_record = {}
                    for src_col, target_param in column_mapping.items():
                        if src_col in row_data:
                            mapped_record[target_param] = row_data[src_col]
                        else:
                            result.warnings.append(
                                f"Row {row_idx+1}: Source column '{src_col}' not found"
                            )

                    if mapped_record:
                        result.imported_data.append(mapped_record)
                        result.records_imported += 1
                    else:
                        result.records_skipped += 1

                except Exception as e:
                    result.errors.append(f"Row {row_idx+1}: {e}")
                    result.records_skipped += 1

            wb.close()
            result.success = result.records_imported > 0

        except Exception as e:
            result.errors.append(f"Import failed: {e}")

        return result

    def get_sheet_names(self, path: Path) -> List[str]:
        """Get list of sheet names in the workbook"""
        if not self._check_openpyxl():
            return []

        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
